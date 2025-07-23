import openpyxl

# Load the original Excel form
form = openpyxl.load_workbook("C:/Users/Admin/Desktop/form.xlsx")
sheet = form.worksheets[0]

# Create a new workbook for processed output
new_form = openpyxl.Workbook()
new_form_sheet = new_form.worksheets[0]

# Set spacing offsets for row and column writing in the new sheet
row_spacing = 3
column_spacing = 2

# Dictionary mapping qualitative answers to numerical indicators
indicators = {
    "Strongly Disagree": 1,
    "Disagree": 2,
    "Agree": 3,
    "Strongly Agree": 4,
    "Extremely High": 5,
    "Very High": 4,
    "Moderately High": 3,
    "Low": 2,
    "Very Low": 1
}

# Column mappings for Table 1 questions in the original sheet
table_1 = {
    "Q1": 8, "Q2": 9, "Q3": 10, "Q4": 11, "Q5": 12,
    "Q6": 13, "Q7": 14, "Q8": 15, "Q9": 16, "Q10": 17
}

# Column mappings for Table 2 - Basic Operating System Function
table_2_basic_operating_system_function = {
    "Q1": 18, "Q2": 19, "Q3": 20, "Q4": 21, "Q5": 22,
    "Q6": 23, "Q7": 24, "Q8": 25, "Q9": 26
}

# Column mappings for Table 2 - Word Processing
table_2_word_processing = {
    "Q1": 27, "Q2": 28, "Q3": 29, "Q4": 30, "Q5": 31,
    "Q6": 32, "Q7": 33, "Q8": 34, "Q9": 35
}

# Column mappings for Table 2 - Spreadsheet
table_2_spreadsheet = {
    "Q1": 36, "Q2": 37, "Q3": 38, "Q4": 39, "Q5": 40,
    "Q6": 41, "Q7": 42, "Q8": 43, "Q9": 44
}

# Column mappings for Table 2 - Multimedia Presentation
table_2_multimedia_presentation = {
    "Q1": 45, "Q2": 46, "Q3": 47, "Q4": 48, "Q5": 49,
    "Q6": 50, "Q7": 51, "Q8": 52, "Q9": 53
}

# Add header for number of respondents
new_form_sheet.cell(3, 1, "No. Of Respondents")

# Loop through each row of the input sheet
for i, row in enumerate(sheet):
    if i == 286:
        break  # Stop after 285 rows (index 0–285)

    if i != 0:  # Skip header row
        # Write respondent number
        new_form_sheet.cell(i + row_spacing, 1, i)

        # Write responses from Table 1 using the indicators mapping
        for j in range(10):
            new_form_sheet.cell(i + row_spacing, j + 1 + column_spacing,
                                indicators[row[table_1[f"Q{j+1}"]].value])

        # Write responses for Table 2 - Basic OS Function
        for j in range(9):
            new_form_sheet.cell(i + row_spacing, j + 14 + column_spacing,
                                indicators[row[table_2_basic_operating_system_function[f"Q{j+1}"]].value])

        # Write responses for Table 2 - Word Processing
        for j in range(9):
            new_form_sheet.cell(i + row_spacing, j + 24 + column_spacing,
                                indicators[row[table_2_word_processing[f"Q{j+1}"]].value])

        # Write responses for Table 2 - Spreadsheet
        for j in range(9):
            new_form_sheet.cell(i + row_spacing, j + 34 + column_spacing,
                                indicators[row[table_2_spreadsheet[f"Q{j+1}"]].value])

        # Write responses for Table 2 - Multimedia Presentation (only Q1 to Q8)
        for j in range(8):  # Q1 to Q8 only (no Q9 in this table)
            new_form_sheet.cell(i + row_spacing, j + 44 + column_spacing,
                                indicators[row[table_2_multimedia_presentation[f"Q{j+1}"]].value])

# Save the new processed workbook
new_form.save("C:/Users/Admin/Desktop/aadfdssa.xlsx")
